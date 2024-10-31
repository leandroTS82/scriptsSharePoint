"""
create_excel.py

Este script lê arquivos CSV de um diretório especificado, processa os dados, 
e os salva em um arquivo Excel com várias abas, além de gerar um resumo de acessos 
dos usuários. 

Dependências:
- pandas: biblioteca para manipulação e análise de dados.
- openpyxl: biblioteca para leitura e escrita de arquivos Excel (XLSX).
  
Como executar:
1. Certifique-se de ter o Python instalado em seu sistema (versão 3.6 ou superior).
2. Instale as dependências necessárias usando pip:
   pip install pandas openpyxl
3. Coloque os arquivos CSV no diretório especificado em 'csv_directory'.
4. Execute o script pelo terminal ou prompt de comando:
   python create_excel.py
5. O arquivo Excel será salvo no diretório 'output_directory' com o nome 
   "Butterfly - Gerenciamento de Acessos.xlsx".
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font

# Defina o diretório onde os arquivos CSV estão localizados
csv_directory = ".\\files\\07csv"
output_directory = ".\\files\\07Excel"
output_file_path = os.path.join(output_directory, "Butterfly - Gerenciamento de Acessos.xlsx")

# Crie o diretório de saída para os arquivos Excel, caso não exista
if not os.path.exists(output_directory):
    os.makedirs(output_directory)

# Crie um objeto ExcelWriter para salvar as várias abas
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    # Itere sobre cada arquivo CSV no diretório
    for csv_file in os.listdir(csv_directory):
        if csv_file.endswith(".csv"):
            # Leia o arquivo CSV em um DataFrame
            csv_path = os.path.join(csv_directory, csv_file)
            df = pd.read_csv(csv_path)

            # Remova a coluna "Login", se existir
            if "Login" in df.columns:
                df = df.drop(columns=["Login"])

            # Extraia o nome do site (sem extensão) para usar como nome da aba
            sheet_name = os.path.splitext(csv_file)[0]
            table_name = sheet_name.replace(" ", "_")  # Substitua espaços por sublinhados

            # Escreva o DataFrame em uma nova aba no arquivo Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Obtenha o objeto do workbook e a aba recém-criada
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Defina o intervalo da tabela
            last_row = len(df) + 1  # +1 para incluir o cabeçalho
            table_range = f"A1:{chr(64 + len(df.columns))}{last_row}"  # Gera o range da tabela (ex: A1:C10)

            # Crie a tabela
            table = Table(displayName=table_name, ref=table_range)  # Use o novo nome da tabela

            # Adicione um estilo à tabela
            style = TableStyleInfo(
                name="TableStyleMedium9", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True,
                showColumnStripes=True
            )
            table.tableStyleInfo = style

            # Adicione a tabela à planilha
            worksheet.add_table(table)

            print(f"Aba '{sheet_name}' criada com sucesso como tabela.")

# Carregar o arquivo existente
try:
    df = pd.read_excel(output_file_path, sheet_name=None, engine='openpyxl')  # Lê todas as planilhas
except Exception as e:
    print(f"Erro ao carregar o arquivo: {e}")

# Criar uma lista única de nomes de usuários
all_users = set()

# Percorre as abas existentes para coletar usuários
for sheet_name, sheet_data in df.items():
    if 'Nome' in sheet_data.columns:  # Verifique se a coluna 'Nome' existe
        users = sheet_data['Nome'].drop_duplicates().tolist()
        all_users.update(users)

# Criar um DataFrame para a nova aba
summary_data = {'Nome': list(all_users)}

# Adicionar uma coluna para cada site
for sheet_name in df.keys():
    summary_data[sheet_name] = [0] * len(all_users)  # Inicializa com 0

# Preencher a tabela com 1s onde o usuário pertence ao site
for sheet_name, sheet_data in df.items():
    if 'Nome' in sheet_data.columns:
        for user in all_users:
            if user in sheet_data['Nome'].values:
                index = summary_data['Nome'].index(user)
                summary_data[sheet_name][index] = 1  # Marca como pertencente

# Criar um DataFrame com os dados resumidos
summary_df = pd.DataFrame(summary_data)

# Adicionar o DataFrame à nova aba como tabela
with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a') as writer:
    summary_df.to_excel(writer, sheet_name='Resumo de Acessos', index=False)

    # Obter o objeto do workbook e a aba recém-criada
    workbook = writer.book
    worksheet = writer.sheets['Resumo de Acessos']

    # Definir o intervalo da tabela
    last_row = len(summary_df) + 1  # +1 para incluir o cabeçalho
    table_range = f"A1:{chr(64 + len(summary_df.columns))}{last_row}"  # Gera o range da tabela

    # Criar a tabela
    table = Table(displayName="ResumoAcessos", ref=table_range)

    # Adicionar um estilo à tabela
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style

    # Adicionar a tabela à planilha
    worksheet.add_table(table)

    # Mover a aba 'Resumo de Acessos' para a primeira posição
    workbook.move_sheet(worksheet, offset=-len(workbook.worksheets) + 1)

    # Adicionar formatação condicional
    # Preenchimento vermelho claro e texto vermelho escuro para 0
    red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
    red_font = Font(color="A00000")

    # Preenchimento verde claro e texto verde escuro para 1
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    green_font = Font(color="005700")

    # Definir regras de formatação condicional
    for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=last_row, max_col=len(summary_df.columns)):
        for cell in row:
            # Adicionar regra para 0
            if cell.value == 0:
                cell.fill = red_fill
                cell.font = red_font
            # Adicionar regra para 1
            elif cell.value == 1:
                cell.fill = green_fill
                cell.font = green_font

print("Aba 'Resumo de Acessos' criada com sucesso como tabela e movida para a primeira posição!")
